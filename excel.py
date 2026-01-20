# STANDALONE EXCEL EXPORT UTILITY
# This script can be run independently to export parking data to Excel

import psycopg
from psycopg.rows import dict_row
from datetime import datetime
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import sys
import os

# Database configuration
DB_CONFIG = {
    'host': 'localhost',
    'dbname': 'parking_system',
    'user': 'postgres',
    'password': 'postgres'
}

def get_db():
    return psycopg.connect(**DB_CONFIG, row_factory=dict_row)

def get_status_color(hours_parked):
    """Determine status color and emoji based on parking duration"""
    if hours_parked >= 24:
        return {'color': 'FF0000', 'emoji': '🔴', 'text': 'Over 24h'}
    elif hours_parked >= 8:
        return {'color': 'FF8C00', 'emoji': '🟡', 'text': '8-24h'}
    else:
        return {'color': '00AA00', 'emoji': '🟢', 'text': 'Under 8h'}

def export_excel_to_file(shift=None, worker_id=None, date_filter=None, output_file=None):
    """
    Export parking data to Excel file

    Args:
        shift (int): Filter by shift number (1-5)
        worker_id (int): Filter by worker ID
        date_filter (str): Date in YYYY-MM-DD format
        output_file (str): Output filename (optional)
    """
    if date_filter is None:
        # Default to today in South Africa timezone
        try:
            tz = pytz.timezone('Africa/Johannesburg')
            date_filter = datetime.now(tz).date().isoformat()
        except:
            date_filter = datetime.now().date().isoformat()

    print(f"Exporting data for date: {date_filter}")

    try:
        conn = get_db()
        cur = conn.cursor()

        # Check if holding columns exist
        cur.execute("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_name = 'cars'
            AND column_name IN ('vessel_id', 'holding_area_id', 'stack_number', 'is_in_holding')
        """)
        existing_columns = [row['column_name'] for row in cur.fetchall()]
        has_holding_columns = len(existing_columns) == 4

        # Query for all cars
        if has_holding_columns:
            query = '''
                SELECT DISTINCT
                    c.car_identifier,
                    c.first_scan_time,
                    c.last_scan_time,
                    c.scan_count,
                    c.status,
                    c.date,
                    c.is_in_holding,
                    u.full_name as worker_name,
                    s.shift_number,
                    EXTRACT(EPOCH FROM (NOW() - c.first_scan_time))/3600 as hours_parked,
                    v.vessel_name,
                    v.vessel_type,
                    ha.area_name as holding_area_name,
                    c.stack_number,
                    (SELECT u2.full_name FROM scans s2
                     JOIN users u2 ON s2.worker_id = u2.user_id
                     WHERE s2.car_id = c.car_id
                     ORDER BY s2.scan_time DESC LIMIT 1) as last_scanned_by,
                    (SELECT s2.scan_time FROM scans s2
                     WHERE s2.car_id = c.car_id
                     ORDER BY s2.scan_time DESC LIMIT 1) as last_scan_time_actual,
                    (SELECT COUNT(*) FROM scans s3 WHERE s3.car_id = c.car_id) as total_scans
                FROM cars c
                JOIN scans s ON c.car_id = s.car_id
                JOIN users u ON s.worker_id = u.user_id
                LEFT JOIN vessels v ON c.vessel_id = v.vessel_id
                LEFT JOIN holding_areas ha ON c.holding_area_id = ha.holding_area_id
                WHERE c.date = %s
            '''
        else:
            query = '''
                SELECT DISTINCT
                    c.car_identifier,
                    c.first_scan_time,
                    c.last_scan_time,
                    c.scan_count,
                    c.status,
                    c.date,
                    FALSE as is_in_holding,
                    u.full_name as worker_name,
                    s.shift_number,
                    EXTRACT(EPOCH FROM (NOW() - c.first_scan_time))/3600 as hours_parked,
                    (SELECT u2.full_name FROM scans s2
                     JOIN users u2 ON s2.worker_id = u2.user_id
                     WHERE s2.car_id = c.car_id
                     ORDER BY s2.scan_time DESC LIMIT 1) as last_scanned_by,
                    (SELECT s2.scan_time FROM scans s2
                     WHERE s2.car_id = c.car_id
                     ORDER BY s2.scan_time DESC LIMIT 1) as last_scan_time_actual,
                    (SELECT COUNT(*) FROM scans s3 WHERE s3.car_id = c.car_id) as total_scans
                FROM cars c
                JOIN scans s ON c.car_id = s.car_id
                JOIN users u ON s.worker_id = u.user_id
                WHERE c.date = %s
            '''

        params = [date_filter]

        if shift:
            query += ' AND s.shift_number = %s'
            params.append(shift)

        if worker_id:
            query += ' AND s.worker_id = %s'
            params.append(worker_id)

        query += ' ORDER BY c.first_scan_time'

        cur.execute(query, params)
        all_data = cur.fetchall()
        conn.close()

        print(f"Found {len(all_data)} records")

        # Separate holding and parked cars
        holding_data = [row for row in all_data if row.get('is_in_holding')]
        parked_data = [row for row in all_data if not row.get('is_in_holding')]

        wb = Workbook()
        try:
            tz = pytz.timezone('Africa/Johannesburg')
            now = datetime.now(tz).replace(tzinfo=None)
        except:
            now = datetime.now()

        # ==================== SHEET 1: HOLDING AREA ====================
        if holding_data:
            ws_holding = wb.active
            ws_holding.title = "Holding Area"

            # Headers for Holding Area
            headers_holding = ['Car ID', 'Vessel', 'Area', 'Unit', 'Worker', 'Time', 'Hours', 'Status', 'Date']
            ws_holding.append(headers_holding)

            # Header styling
            header_fill = PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for cell in ws_holding[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add holding area data
            for row in holding_data:
                hours_parked = row['hours_parked']
                status_info = get_status_color(hours_parked)

                scan_time = row['last_scan_time'].strftime('%I:%M %p')
                vessel_info = f"{row.get('vessel_name', '-')} ({row.get('vessel_type', '-')})" if row.get('vessel_name') else '-'
                status_display = f"{status_info['emoji']} {status_info['text']}"

                row_data = [
                    row['car_identifier'],
                    vessel_info,
                    row.get('holding_area_name', '-') or '-',
                    row.get('stack_number', '-') or '-',
                    row['worker_name'],
                    scan_time,
                    f"{hours_parked:.1f}h",
                    status_display,
                    str(row['date'])
                ]
                ws_holding.append(row_data)

            # Color code status column
            for row_idx in range(2, ws_holding.max_row + 1):
                status_cell = ws_holding.cell(row=row_idx, column=8)
                if status_cell.value and '🔴' in str(status_cell.value):
                    status_cell.font = Font(color='FF0000', bold=True, size=12)
                elif status_cell.value and '🟡' in str(status_cell.value):
                    status_cell.font = Font(color='FF8C00', bold=True, size=12)
                elif status_cell.value and '🟢' in str(status_cell.value):
                    status_cell.font = Font(color='00AA00', bold=True, size=12)

            # Auto-adjust columns
            for col in ws_holding.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_holding.column_dimensions[column].width = min(max_length + 2, 50)

        # ==================== SHEET 2: PARKED VEHICLES ====================
        if parked_data:
            if not holding_data:
                ws_parked = wb.active
                ws_parked.title = "Parked Vehicles"
            else:
                ws_parked = wb.create_sheet("Parked Vehicles")

            # Headers for Parked Vehicles
            headers_parked = ['Car ID', 'First Scan', 'Last Scan', 'Worker', 'Last Scanned By',
                             'Time Difference', 'Scans', 'Hours', 'Status', 'Date']
            ws_parked.append(headers_parked)

            # Header styling
            header_fill = PatternFill(start_color='6366f1', end_color='6366f1', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')

            for cell in ws_parked[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add parked vehicles data
            for row in parked_data:
                hours_parked = row['hours_parked']
                status_info = get_status_color(hours_parked)

                first_scan_time = row['first_scan_time'].strftime('%I:%M %p')
                last_scan_time = row['last_scan_time'].strftime('%I:%M %p')

                # Time difference between first and last scan
                time_diff = row['last_scan_time'] - row['first_scan_time']
                hours = int(time_diff.total_seconds() // 3600)
                minutes = int((time_diff.total_seconds() % 3600) // 60)
                time_diff_str = f"{hours}h {minutes}m"

                # Last Scanned By (only for repeated scans)
                last_scanned_by = row.get('last_scanned_by', 'Unknown')
                total_scans = row.get('total_scans', 1)

                if total_scans > 1 and row.get('last_scan_time_actual'):
                    time_since_scan = (now - row['last_scan_time_actual']).total_seconds() / 3600
                    if time_since_scan < 1:
                        time_ago_str = f"{int(time_since_scan * 60)} min ago"
                    else:
                        time_ago_str = f"{int(time_since_scan)}h ago"
                    last_scanned_display = f"{time_ago_str} by {last_scanned_by}"
                else:
                    last_scanned_display = "-"  # Not a repeated scan

                status_display = f"{status_info['emoji']} {status_info['text']}"

                row_data = [
                    row['car_identifier'],
                    first_scan_time,
                    last_scan_time,
                    row['worker_name'],
                    last_scanned_display,
                    time_diff_str,
                    f"{row['scan_count']}x",
                    f"{hours_parked:.1f}h",
                    status_display,
                    str(row['date'])
                ]
                ws_parked.append(row_data)

            # Color code status column
            for row_idx in range(2, ws_parked.max_row + 1):
                status_cell = ws_parked.cell(row=row_idx, column=9)
                if status_cell.value and '🔴' in str(status_cell.value):
                    status_cell.font = Font(color='FF0000', bold=True, size=12)
                elif status_cell.value and '🟡' in str(status_cell.value):
                    status_cell.font = Font(color='FF8C00', bold=True, size=12)
                elif status_cell.value and '🟢' in str(status_cell.value):
                    status_cell.font = Font(color='00AA00', bold=True, size=12)

            # Auto-adjust columns
            for col in ws_parked.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws_parked.column_dimensions[column].width = min(max_length + 2, 50)

        # If no data, create empty sheet
        if not holding_data and not parked_data:
            ws = wb.active
            ws.title = "No Data"
            ws['A1'] = "No vehicle data for selected date"

        # Generate filename
        filename = f"parking_report_{date_filter}"
        if shift:
            filename += f"_shift{shift}"
        if worker_id:
            filename += f"_worker{worker_id}"
        filename += ".xlsx"

        if output_file:
            filename = output_file

        # Save file
        wb.save(filename)
        print(f"Excel file saved as: {filename}")
        return filename

    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        import traceback
        traceback.print_exc()
        return None

# Remove the Flask route decorator and function - this is now a standalone utility
# @app.route('/api/export', methods=['GET'])
# @token_required
# def export_excel(current_user):
    
    # Check if holding columns exist
    cur.execute("""
        SELECT column_name 
        FROM information_schema.columns 
        WHERE table_name = 'cars' 
        AND column_name IN ('vessel_id', 'holding_area_id', 'stack_number', 'is_in_holding')
    """)
    existing_columns = [row['column_name'] for row in cur.fetchall()]
    has_holding_columns = len(existing_columns) == 4
    
    # Query for all cars
    if has_holding_columns:
        query = '''
            SELECT DISTINCT
                c.car_identifier, 
                c.first_scan_time, 
                c.last_scan_time, 
                c.scan_count,
                c.status, 
                c.date,
                c.is_in_holding,
                u.full_name as worker_name, 
                s.shift_number,
                EXTRACT(EPOCH FROM (NOW() - c.first_scan_time))/3600 as hours_parked,
                v.vessel_name, 
                v.vessel_type,
                ha.area_name as holding_area_name,
                c.stack_number,
                (SELECT u2.full_name FROM scans s2 
                 JOIN users u2 ON s2.worker_id = u2.user_id 
                 WHERE s2.car_id = c.car_id 
                 ORDER BY s2.scan_time DESC LIMIT 1) as last_scanned_by,
                (SELECT s2.scan_time FROM scans s2 
                 WHERE s2.car_id = c.car_id 
                 ORDER BY s2.scan_time DESC LIMIT 1) as last_scan_time_actual,
                (SELECT COUNT(*) FROM scans s3 WHERE s3.car_id = c.car_id) as total_scans
            FROM cars c
            JOIN scans s ON c.car_id = s.car_id
            JOIN users u ON s.worker_id = u.user_id
            LEFT JOIN vessels v ON c.vessel_id = v.vessel_id
            LEFT JOIN holding_areas ha ON c.holding_area_id = ha.holding_area_id
            WHERE c.date = %s
        '''
    else:
        query = '''
            SELECT DISTINCT
                c.car_identifier, 
                c.first_scan_time, 
                c.last_scan_time, 
                c.scan_count,
                c.status, 
                c.date,
                FALSE as is_in_holding,
                u.full_name as worker_name, 
                s.shift_number,
                EXTRACT(EPOCH FROM (NOW() - c.first_scan_time))/3600 as hours_parked,
                (SELECT u2.full_name FROM scans s2 
                 JOIN users u2 ON s2.worker_id = u2.user_id 
                 WHERE s2.car_id = c.car_id 
                 ORDER BY s2.scan_time DESC LIMIT 1) as last_scanned_by,
                (SELECT s2.scan_time FROM scans s2 
                 WHERE s2.car_id = c.car_id 
                 ORDER BY s2.scan_time DESC LIMIT 1) as last_scan_time_actual,
                (SELECT COUNT(*) FROM scans s3 WHERE s3.car_id = c.car_id) as total_scans
            FROM cars c
            JOIN scans s ON c.car_id = s.car_id
            JOIN users u ON s.worker_id = u.user_id
            WHERE c.date = %s
        '''
    
    params = [date_filter]
    
    if current_user['role'] == 'worker':
        query += ' AND s.worker_id = %s'
        params.append(current_user['user_id'])
    elif shift:
        query += ' AND s.shift_number = %s'
        params.append(shift)
    
    if worker_id:
        query += ' AND s.worker_id = %s'
        params.append(worker_id)
    
    query += ' ORDER BY c.first_scan_time'
    
    cur.execute(query, params)
    all_data = cur.fetchall()
    conn.close()
    
    # Separate holding and parked cars
    holding_data = [row for row in all_data if row.get('is_in_holding')]
    parked_data = [row for row in all_data if not row.get('is_in_holding')]
    
    wb = Workbook()
    now = datetime.now(pytz.timezone('Africa/Johannesburg')).replace(tzinfo=None)
    
    # ==================== SHEET 1: HOLDING AREA ====================
    if holding_data:
        ws_holding = wb.active
        ws_holding.title = "Holding Area"
        
        # Headers for Holding Area
        headers_holding = ['Car ID', 'Vessel', 'Area', 'Unit', 'Worker', 'Time', 'Hours', 'Status', 'Date']
        ws_holding.append(headers_holding)
        
        # Header styling
        header_fill = PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws_holding[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add holding area data
        for row in holding_data:
            hours_parked = row['hours_parked']
            status_info = get_status_color(hours_parked)
            
            scan_time = row['last_scan_time'].strftime('%I:%M %p')
            vessel_info = f"{row.get('vessel_name', '-')} ({row.get('vessel_type', '-')})" if row.get('vessel_name') else '-'
            status_display = f"{status_info['emoji']} {status_info['text']}"
            
            row_data = [
                row['car_identifier'],
                vessel_info,
                row.get('holding_area_name', '-') or '-',
                row.get('stack_number', '-') or '-',
                row['worker_name'],
                scan_time,
                f"{hours_parked:.1f}h",
                status_display,
                str(row['date'])
            ]
            ws_holding.append(row_data)
        
        # Color code status column
        for row_idx in range(2, ws_holding.max_row + 1):
            status_cell = ws_holding.cell(row=row_idx, column=8)
            if status_cell.value and '🔴' in str(status_cell.value):
                status_cell.font = Font(color='FF0000', bold=True, size=12)
            elif status_cell.value and '🟡' in str(status_cell.value):
                status_cell.font = Font(color='FF8C00', bold=True, size=12)
            elif status_cell.value and '🟢' in str(status_cell.value):
                status_cell.font = Font(color='00AA00', bold=True, size=12)
        
        # Auto-adjust columns
        for col in ws_holding.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_holding.column_dimensions[column].width = min(max_length + 2, 50)
    
    # ==================== SHEET 2: PARKED VEHICLES ====================
    if parked_data:
        if not holding_data:
            ws_parked = wb.active
            ws_parked.title = "Parked Vehicles"
        else:
            ws_parked = wb.create_sheet("Parked Vehicles")
        
        # Headers for Parked Vehicles
        headers_parked = ['Car ID', 'First Scan', 'Last Scan', 'Worker', 'Last Scanned By', 
                         'Time Difference', 'Scans', 'Hours', 'Status', 'Date']
        ws_parked.append(headers_parked)
        
        # Header styling
        header_fill = PatternFill(start_color='6366f1', end_color='6366f1', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws_parked[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Add parked vehicles data
        for row in parked_data:
            hours_parked = row['hours_parked']
            status_info = get_status_color(hours_parked)
            
            first_scan_time = row['first_scan_time'].strftime('%I:%M %p')
            last_scan_time = row['last_scan_time'].strftime('%I:%M %p')
            
            # Time difference between first and last scan
            time_diff = row['last_scan_time'] - row['first_scan_time']
            hours = int(time_diff.total_seconds() // 3600)
            minutes = int((time_diff.total_seconds() % 3600) // 60)
            time_diff_str = f"{hours}h {minutes}m"
            
            # Last Scanned By (only for repeated scans)
            last_scanned_by = row.get('last_scanned_by', 'Unknown')
            total_scans = row.get('total_scans', 1)
            
            if total_scans > 1 and row.get('last_scan_time_actual'):
                time_since_scan = (now - row['last_scan_time_actual']).total_seconds() / 3600
                if time_since_scan < 1:
                    time_ago_str = f"{int(time_since_scan * 60)} min ago"
                else:
                    time_ago_str = f"{int(time_since_scan)}h ago"
                last_scanned_display = f"{time_ago_str} by {last_scanned_by}"
            else:
                last_scanned_display = "-"  # Not a repeated scan
            
            status_display = f"{status_info['emoji']} {status_info['text']}"
            
            row_data = [
                row['car_identifier'],
                first_scan_time,
                last_scan_time,
                row['worker_name'],
                last_scanned_display,
                time_diff_str,
                f"{row['scan_count']}x",
                f"{hours_parked:.1f}h",
                status_display,
                str(row['date'])
            ]
            ws_parked.append(row_data)
        
        # Color code status column
        for row_idx in range(2, ws_parked.max_row + 1):
            status_cell = ws_parked.cell(row=row_idx, column=9)
            if status_cell.value and '🔴' in str(status_cell.value):
                status_cell.font = Font(color='FF0000', bold=True, size=12)
            elif status_cell.value and '🟡' in str(status_cell.value):
                status_cell.font = Font(color='FF8C00', bold=True, size=12)
            elif status_cell.value and '🟢' in str(status_cell.value):
                status_cell.font = Font(color='00AA00', bold=True, size=12)
        
        # Auto-adjust columns
        for col in ws_parked.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws_parked.column_dimensions[column].width = min(max_length + 2, 50)
    
    # If no data, create empty sheet
    if not holding_data and not parked_data:
        ws = wb.active
        ws.title = "No Data"
        ws['A1'] = "No vehicle data for selected date"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"parking_report_{date_filter}"
    if shift:
        filename += f"_shift{shift}"
    if worker_id:
        filename += f"_worker{worker_id}"
    filename += ".xlsx"

    # Save file
    wb.save(filename)
    print(f"Excel file saved as: {filename}")
    return filename

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description='Export parking data to Excel')
    parser.add_argument('--shift', type=int, help='Filter by shift number (1-5)')
    parser.add_argument('--worker-id', type=int, help='Filter by worker ID')
    parser.add_argument('--date', type=str, help='Date in YYYY-MM-DD format (default: today)')
    parser.add_argument('--output', type=str, help='Output filename (default: auto-generated)')

    args = parser.parse_args()

    print("Parking Data Excel Export Utility")
    print("=" * 40)

    result = export_excel_to_file(
        shift=args.shift,
        worker_id=args.worker_id,
        date_filter=args.date,
        output_file=args.output
    )

    if result:
        print(f"\n✅ Export completed successfully!")
        print(f"📁 File saved: {result}")
    else:
        print("\n❌ Export failed!")
        sys.exit(1)