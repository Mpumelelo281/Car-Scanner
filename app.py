from flask import Flask, request, jsonify, send_file, render_template
from flask_cors import CORS
from datetime import datetime, timedelta
import pytz
import psycopg
from psycopg.rows import dict_row
import bcrypt
import jwt
from functools import wraps
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

app = Flask(__name__)
CORS(app)

app.config['SECRET_KEY'] = 'parking-system-secret-key-2025'
DB_CONFIG = {
    'host': 'localhost',
    'dbname': 'parking_system',
    'user': 'postgres',
    'password': 'postgres'
}

# South Africa Timezone
SOUTH_AFRICA_TZ = pytz.timezone('Africa/Johannesburg')

def get_current_time():
    """Get current time in South Africa timezone"""
    return datetime.now(SOUTH_AFRICA_TZ)

def get_db():
    return psycopg.connect(**DB_CONFIG, row_factory=dict_row)

SHIFTS = {
    1: {'start': 6, 'end': 18, 'name': '6AM-6PM (Day Shift)'},
    2: {'start': 18, 'end': 6, 'name': '6PM-6AM (Night Shift)'}
}

def get_current_shift():
    """Get current shift based on South Africa time"""
    hour = get_current_time().hour
    return 1 if 6 <= hour < 18 else 2

def get_status_color(hours_parked):
    """Get status based on hours: green < 4h, amber 4-12h, red 12h+"""
    if hours_parked < 4:
        return {'emoji': '🟢', 'status': 'green', 'text': 'Normal'}
    elif hours_parked < 12:
        return {'emoji': '🟡', 'status': 'amber', 'text': 'Warning'}
    else:
        return {'emoji': '🔴', 'status': 'red', 'text': 'Overdue'}

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'error': 'Token missing'}), 401
        try:
            token = token.split(' ')[1] if ' ' in token else token
            data = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
            current_user = data
        except Exception as e:
            print(f"Token decode error: {e}")
            return jsonify({'error': 'Invalid token'}), 401
        return f(current_user, *args, **kwargs)
    return decorated

def role_required(roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(current_user, *args, **kwargs):
            if current_user['role'] not in roles:
                return jsonify({'error': 'Unauthorized'}), 403
            return f(current_user, *args, **kwargs)
        return decorated_function
    return decorator

# Routes that were causing 404 errors
@app.route('/favicon.ico')
def favicon():
    return '', 204

@app.route('/.well-known/<path:path>')
def well_known(path):
    return '', 204

@app.route('/')
def index():
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/health', methods=['GET'])
def health():
    try:
        conn = get_db()
        conn.close()
        return jsonify({'status': 'healthy', 'db': 'connected'}), 200
    except Exception as e:
        return jsonify({'status': 'unhealthy', 'error': str(e)}), 500

# AUTH - FIXED
@app.route('/api/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({'error': 'Username and password required'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT * FROM users WHERE username = %s AND is_active = TRUE', (username,))
        user = cur.fetchone()
        conn.close()
        
        if not user:
            return jsonify({'error': 'Invalid credentials'}), 401
        
        if not user['password_hash']:
            return jsonify({'error': 'User account needs password setup'}), 401
        
        password_hash_bytes = user['password_hash']
        if isinstance(password_hash_bytes, str):
            password_hash_bytes = password_hash_bytes.encode('utf-8')
        
        if bcrypt.checkpw(password.encode(), password_hash_bytes):
            assigned_shift = user['assigned_shift'] or get_current_shift()
            
            token = jwt.encode({
                'user_id': user['user_id'],
                'username': user['username'],
                'role': user['role'],
                'full_name': user['full_name'] or user['username'],
                'assigned_shift': assigned_shift,
                'supervisor_id': user['supervisor_id']
            }, app.config['SECRET_KEY'], algorithm='HS256')
            
            return jsonify({
                'token': token,
                'user': {
                    'user_id': user['user_id'],
                    'username': user['username'],
                    'role': user['role'],
                    'full_name': user['full_name'] or user['username'],
                    'assigned_shift': assigned_shift
                }
            }), 200
        else:
            return jsonify({'error': 'Invalid credentials'}), 401
            
    except Exception as e:
        print(f"Login error: {e}")
        return jsonify({'error': 'Server error during login'}), 500

# HOLDING AREAS
@app.route('/api/holding-areas', methods=['GET'])
@token_required
def get_holding_areas(current_user):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = 'holding_areas'
            )
        """)
        if not cur.fetchone()['exists']:
            conn.close()
            return jsonify([
                {'holding_area_id': 1, 'area_name': 'Holding Area A'},
                {'holding_area_id': 2, 'area_name': 'Holding Area B'},
                {'holding_area_id': 3, 'area_name': 'Holding Area C'}
            ])
        
        cur.execute('SELECT * FROM holding_areas WHERE is_active = TRUE ORDER BY area_name')
        areas = cur.fetchall()
        conn.close()
        return jsonify([dict(a) for a in areas])
    except Exception as e:
        print(f"Error loading holding areas: {e}")
        return jsonify([])

# VESSELS
@app.route('/api/vessels', methods=['GET'])
@token_required
def get_vessels(current_user):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('SELECT * FROM vessels WHERE is_active = TRUE ORDER BY arrival_date DESC')
        vessels = cur.fetchall()
        conn.close()
        return jsonify([dict(v) for v in vessels])
    except:
        return jsonify([])

@app.route('/api/vessels', methods=['POST'])
@token_required
def create_vessel(current_user):
    try:
        data = request.json
        vessel_name = data.get('vessel_name', '').strip()
        vessel_type = data.get('vessel_type', 'ship')
        arrival_date = data.get('arrival_date')
        
        if not vessel_name:
            return jsonify({'error': 'Vessel name required'}), 400
        
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''
            INSERT INTO vessels (vessel_name, vessel_type, arrival_date)
            VALUES (%s, %s, %s) RETURNING vessel_id
        ''', (vessel_name, vessel_type, arrival_date))
        vessel_id = cur.fetchone()['vessel_id']
        conn.commit()
        conn.close()
        return jsonify({'message': 'Vessel created', 'vessel_id': vessel_id}), 201
    except Exception as e:
        print(f"Error creating vessel: {e}")
        return jsonify({'error': str(e)}), 500

# USERS
@app.route('/api/users', methods=['GET'])
@token_required
def get_users(current_user):
    conn = get_db()
    cur = conn.cursor()
    
    if current_user['role'] == 'supervisor':
        cur.execute('''
            SELECT u.*, s.full_name as supervisor_name
            FROM users u
            LEFT JOIN users s ON u.supervisor_id = s.user_id
            WHERE u.supervisor_id = %s AND u.role = 'worker'
            ORDER BY u.full_name
        ''', (current_user['user_id'],))
    else:
        cur.execute('''
            SELECT u.*, s.full_name as supervisor_name
            FROM users u
            LEFT JOIN users s ON u.supervisor_id = s.user_id
            ORDER BY u.role, u.full_name
        ''')
    
    users = cur.fetchall()
    conn.close()
    return jsonify([dict(u) for u in users])

@app.route('/api/users', methods=['POST'])
@token_required
@role_required(['admin'])
def create_user(current_user):
    try:
        data = request.json
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        full_name = data.get('full_name', '').strip()
        role = data.get('role', '').strip()
        
        if not username or not password or not role:
            return jsonify({'error': 'Missing required fields'}), 400
        
        password_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
        
        assigned_shift = data.get('assigned_shift')
        if assigned_shift in [None, '', 'null']:
            assigned_shift = 1 if role == 'worker' else None
        else:
            assigned_shift = int(assigned_shift)
        
        supervisor_id = data.get('supervisor_id')
        if supervisor_id in [None, '', 'null']:
            supervisor_id = None
        else:
            supervisor_id = int(supervisor_id)
        
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''
            INSERT INTO users (username, password_hash, role, full_name, assigned_shift, supervisor_id)
            VALUES (%s, %s, %s, %s, %s, %s) RETURNING user_id
        ''', (username, password_hash, role, full_name, assigned_shift, supervisor_id))
        
        user_id = cur.fetchone()['user_id']
        conn.commit()
        conn.close()
        return jsonify({'message': 'User created', 'user_id': user_id}), 201
        
    except psycopg.errors.UniqueViolation:
        return jsonify({'error': 'Username already exists'}), 400
    except Exception as e:
        print(f"Error creating user: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@token_required
@role_required(['admin'])
def delete_user(current_user, user_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute('UPDATE users SET is_active = FALSE WHERE user_id = %s', (user_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'User deactivated'})

# WORKER PROFILE
@app.route('/api/workers/<int:worker_id>/profile', methods=['GET'])
@token_required
def get_worker_profile(current_user, worker_id):
    try:
        conn = get_db()
        cur = conn.cursor()
        
        cur.execute('SELECT * FROM users WHERE user_id = %s AND role = \'worker\'', (worker_id,))
        worker = cur.fetchone()
        
        if not worker:
            conn.close()
            return jsonify({'error': 'Worker not found'}), 404
        
        today = get_current_time().date()
        
        cur.execute('SELECT COUNT(*) as today_scans FROM scans WHERE worker_id = %s AND date = %s', (worker_id, today))
        today_stats = cur.fetchone()
        
        cur.execute('SELECT COUNT(*) as week_scans FROM scans WHERE worker_id = %s AND date >= CURRENT_DATE - INTERVAL \'7 days\'', (worker_id,))
        week_stats = cur.fetchone()
        
        cur.execute('SELECT COUNT(DISTINCT car_id) as unique_cars FROM scans WHERE worker_id = %s', (worker_id,))
        car_stats = cur.fetchone()
        
        cur.execute('''
            SELECT c.car_identifier, s.scan_time, s.shift_number
            FROM scans s
            JOIN cars c ON s.car_id = c.car_id
            WHERE s.worker_id = %s
            ORDER BY s.scan_time DESC
            LIMIT 10
        ''', (worker_id,))
        recent_scans = cur.fetchall()
        
        conn.close()
        
        return jsonify({
            'worker': dict(worker),
            'stats': {
                'today_scans': today_stats['today_scans'],
                'week_scans': week_stats['week_scans'],
                'total_scans': worker.get('total_scans') or 0,
                'unique_cars': car_stats['unique_cars']
            },
            'recent_activity': [dict(s) for s in recent_scans]
        })
    except Exception as e:
        print(f"Error: {e}")
        return jsonify({'error': str(e)}), 500

# SCANNING - FIXED to use South Africa time
@app.route('/api/scan', methods=['POST'])
@token_required
def scan_car(current_user):
    try:
        data = request.json
        car_identifier = data.get('car_identifier', '').strip().upper()
        
        vessel_id = data.get('vessel_id')
        holding_area_id = data.get('holding_area_id')
        stack_number = data.get('stack_number', '').strip()
        is_in_holding = data.get('is_in_holding', False)
        
        if not car_identifier:
            return jsonify({'error': 'Car identifier required'}), 400
        
        user_id = current_user.get('user_id')
        
        conn = get_db()
        cur = conn.cursor()
        
        # Use South Africa time with timezone info (don't strip tzinfo)
        now = get_current_time()
        today = now.date()
        
        # Check if worker has assigned shift
        user_role = current_user.get('role')
        assigned_shift = current_user.get('assigned_shift')
        current_hour = now.hour
        
        # Only enforce shift restriction for workers (not supervisors or admins)
        if user_role == 'worker' and assigned_shift:
            shift_info = SHIFTS[assigned_shift]
            shift_start = shift_info['start']
            shift_end = shift_info['end']
            
            # Check if current time is within worker's assigned shift
            if assigned_shift == 1:  # Day shift 6AM-6PM
                is_on_shift = shift_start <= current_hour < shift_end
            else:  # Night shift 6PM-6AM (spans midnight)
                is_on_shift = current_hour >= shift_start or current_hour < shift_end
            
            if not is_on_shift:
                shift_name = shift_info['name']
                return jsonify({
                    'error': 'Outside working hours',
                    'message': f'You cannot scan outside your shift. Your shift ({shift_name}) starts at {shift_start}:00',
                    'assigned_shift': assigned_shift,
                    'shift_info': shift_name,
                    'current_hour': current_hour
                }), 403
        
        shift_number = current_user.get('assigned_shift') or get_current_shift()
        
        # Check if car exists
        cur.execute('SELECT * FROM cars WHERE car_identifier = %s AND is_active = TRUE', (car_identifier,))
        car = cur.fetchone()
        
        if car:
            car_id = car['car_id']
            scan_count = car['scan_count'] + 1
            hours_parked = (now - car['first_scan_time']).total_seconds() / 3600
            status_info = get_status_color(hours_parked)
            status = status_info['status']
            
            # Update with holding info if provided
            if is_in_holding:
                cur.execute('''
                    UPDATE cars SET last_scan_time = %s, scan_count = %s, status = %s,
                           vessel_id = %s, holding_area_id = %s, stack_number = %s, is_in_holding = %s
                    WHERE car_id = %s
                ''', (now, scan_count, status, vessel_id, holding_area_id, stack_number, is_in_holding, car_id))
            else:
                cur.execute('''
                    UPDATE cars SET last_scan_time = %s, scan_count = %s, status = %s
                    WHERE car_id = %s
                ''', (now, scan_count, status, car_id))
        else:
            # Create new car
            status = 'green'
            cur.execute('''
                INSERT INTO cars (car_identifier, first_scan_time, last_scan_time, scan_count, status, date,
                                  vessel_id, holding_area_id, stack_number, is_in_holding)
                VALUES (%s, %s, %s, 1, %s, %s, %s, %s, %s, %s) RETURNING car_id
            ''', (car_identifier, now, now, status, today, vessel_id, holding_area_id, stack_number, is_in_holding))
            car_id = cur.fetchone()['car_id']
        
        # Insert scan record
        cur.execute('''
            INSERT INTO scans (car_id, worker_id, scan_time, shift_number, date)
            VALUES (%s, %s, %s, %s, %s)
        ''', (car_id, user_id, now, shift_number, today))
        
        conn.commit()
        
        # Get updated car info
        cur.execute('''
            SELECT c.*, u.full_name as last_worker,
                   v.vessel_name, v.vessel_type,
                   ha.area_name as holding_area_name
            FROM cars c
            LEFT JOIN scans s ON c.car_id = s.car_id AND s.scan_time = c.last_scan_time
            LEFT JOIN users u ON s.worker_id = u.user_id
            LEFT JOIN vessels v ON c.vessel_id = v.vessel_id
            LEFT JOIN holding_areas ha ON c.holding_area_id = ha.holding_area_id
            WHERE c.car_id = %s
        ''', (car_id,))
        updated_car = cur.fetchone()
        
        # Get previous scans
        cur.execute('''
            SELECT s.scan_time, u.full_name as worker_name, s.shift_number
            FROM scans s
            JOIN users u ON s.worker_id = u.user_id
            WHERE s.car_id = %s AND s.worker_id != %s
            ORDER BY s.scan_time DESC
            LIMIT 3
        ''', (car_id, user_id))
        previous_scans = cur.fetchall()
        
        conn.close()
        
        scan_history = []
        for scan in previous_scans:
            time_ago = (now - scan['scan_time']).total_seconds() / 3600
            time_str = f"{int(time_ago * 60)} min ago" if time_ago < 1 else f"{int(time_ago)}h ago"
            scan_history.append({
                'worker': scan['worker_name'],
                'shift': scan['shift_number'],
                'time_ago': time_str
            })
        
        return jsonify({
            'message': 'Scan recorded successfully', 
            'car': dict(updated_car),
            'previous_scans': scan_history,
            'is_new': len(previous_scans) == 0 and car is None
        })
        
    except Exception as e:
        print(f"Scan error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# GET CARS - FIXED
@app.route('/api/cars', methods=['GET'])
@token_required
def get_cars(current_user):
    try:
        shift = request.args.get('shift', type=int)
        date_filter = request.args.get('date', get_current_time().date().isoformat())
        status_filter = request.args.get('status')
        holding_only = request.args.get('holding_only', 'false').lower() == 'true'
        
        conn = get_db()
        cur = conn.cursor()
        
        base_query = '''
            SELECT DISTINCT c.*, 
                   (SELECT u.full_name FROM scans s 
                    JOIN users u ON s.worker_id = u.user_id 
                    WHERE s.car_id = c.car_id 
                    ORDER BY s.scan_time DESC LIMIT 1) as last_worker,
                   (SELECT s.worker_id FROM scans s 
                    WHERE s.car_id = c.car_id 
                    ORDER BY s.scan_time DESC LIMIT 1) as last_worker_id,
                   v.vessel_name, v.vessel_type,
                   ha.area_name as holding_area_name
            FROM cars c
            LEFT JOIN vessels v ON c.vessel_id = v.vessel_id
            LEFT JOIN holding_areas ha ON c.holding_area_id = ha.holding_area_id
            WHERE c.date = %s AND c.is_active = TRUE
        '''
        
        params = [date_filter]
        
        if holding_only:
            base_query += ' AND c.is_in_holding = TRUE'
        
        if current_user['role'] == 'worker':
            base_query += ' AND EXISTS (SELECT 1 FROM scans s WHERE s.car_id = c.car_id AND s.worker_id = %s)'
            params.append(current_user['user_id'])
        elif shift:
            base_query += ' AND EXISTS (SELECT 1 FROM scans s WHERE s.car_id = c.car_id AND s.shift_number = %s)'
            params.append(shift)
        
        if status_filter:
            base_query += ' AND c.status = %s'
            params.append(status_filter)
        
        base_query += ' ORDER BY c.last_scan_time DESC'
        
        cur.execute(base_query, params)
        cars = cur.fetchall()
        conn.close()
        return jsonify([dict(c) for c in cars])
    except Exception as e:
        print(f"Error getting cars: {e}")
        return jsonify({'error': str(e)}), 500

# DASHBOARD
@app.route('/api/dashboard', methods=['GET'])
@token_required
def get_dashboard(current_user):
    try:
        conn = get_db()
        cur = conn.cursor()
        today = get_current_time().date()
        
        if current_user['role'] == 'worker':
            cur.execute('''
                SELECT 
                    COUNT(DISTINCT c.car_id) as total_cars,
                    COUNT(DISTINCT CASE WHEN c.status = 'red' THEN c.car_id END) as overdue_cars,
                    COUNT(DISTINCT CASE WHEN c.status = 'amber' THEN c.car_id END) as warning_cars,
                    COUNT(DISTINCT CASE WHEN c.status = 'green' THEN c.car_id END) as active_cars
                FROM cars c
                JOIN scans s ON c.car_id = s.car_id
                WHERE s.date = %s AND s.worker_id = %s
            ''', (today, current_user['user_id']))
        else:
            cur.execute('''
                SELECT 
                    COUNT(DISTINCT car_id) as total_cars,
                    COUNT(DISTINCT CASE WHEN status = 'red' THEN car_id END) as overdue_cars,
                    COUNT(DISTINCT CASE WHEN status = 'amber' THEN car_id END) as warning_cars,
                    COUNT(DISTINCT CASE WHEN status = 'green' THEN car_id END) as active_cars
                FROM cars WHERE date = %s AND is_active = TRUE
            ''', (today,))
        
        stats = cur.fetchone()
        
        cur.execute('SELECT COUNT(*) as count FROM users WHERE is_active = TRUE AND role = %s', ('worker',))
        worker_stats = cur.fetchone()
        
        conn.close()
        return jsonify({**dict(stats), 'active_workers': worker_stats['count']})
    except Exception as e:
        print(f"Dashboard error: {e}")
        return jsonify({'error': str(e)}), 500

# EXCEL EXPORT - FIXED with proper auth
@app.route('/api/export', methods=['GET'])
@token_required
def export_excel(current_user):
    shift = request.args.get('shift', type=int)
    worker_id = request.args.get('worker_id', type=int)
    date_filter = request.args.get('date', get_current_time().date().isoformat())
    
    conn = get_db()
    cur = conn.cursor()
    
    query = '''
        SELECT c.car_identifier, c.first_scan_time, c.last_scan_time, c.scan_count,
               c.status, c.date, u.full_name as worker_name, s.shift_number,
               EXTRACT(EPOCH FROM (c.last_scan_time - c.first_scan_time))/3600 as hours_parked
        FROM cars c
        JOIN scans s ON c.car_id = s.car_id
        JOIN users u ON s.worker_id = u.user_id
        WHERE c.date = %s AND (c.is_in_holding IS NULL OR c.is_in_holding = FALSE)
    '''
    
    params = [date_filter]
    
    if current_user['role'] == 'worker':
        query += ' AND s.worker_id = %s'
        params.append(current_user['user_id'])
    elif shift:
        query += ' AND s.shift_number = %s'
        params.append(shift)
    
    if worker_id:
        query += ' AND s.worker_id = %s'
        params.append(worker_id)
    
    query += ' ORDER BY c.first_scan_time'
    
    cur.execute(query, params)
    data = cur.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Parking Report"
    
    headers = ['Car ID', 'First Scan', 'Last Scan', 'Time Difference', 'Scans', 
               'Hours Parked', 'Status', 'Flag', 'Worker', 'Shift', 'Date']
    ws.append(headers)
    
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row in data:
        hours_parked = row['hours_parked']
        status_info = get_status_color(hours_parked)
        
        first_scan_time = row['first_scan_time'].strftime('%I:%M %p')
        last_scan_time = row['last_scan_time'].strftime('%I:%M %p')
        
        time_diff = row['last_scan_time'] - row['first_scan_time']
        hours = int(time_diff.total_seconds() // 3600)
        minutes = int((time_diff.total_seconds() % 3600) // 60)
        time_diff_str = f"{hours}h {minutes}m"
        
        ws.append([
            row['car_identifier'],
            first_scan_time,
            last_scan_time,
            time_diff_str,
            f"{row['scan_count']}x",
            f"{hours_parked:.1f}h",
            status_info['text'],
            status_info['emoji'],
            row['worker_name'],
            f"Shift {row['shift_number']}",
            str(row['date'])
        ])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"parking_report_{date_filter}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# HOLDING AREA EXCEL EXPORT - FIXED
@app.route('/api/export/holding', methods=['GET'])
@token_required
def export_holding_excel(current_user):
    shift = request.args.get('shift', type=int)
    date_filter = request.args.get('date', get_current_time().date().isoformat())
    
    conn = get_db()
    cur = conn.cursor()
    
    query = '''
        SELECT DISTINCT
            c.car_identifier, 
            c.first_scan_time,
            c.last_scan_time, 
            u.full_name as worker_name, 
            EXTRACT(EPOCH FROM (c.last_scan_time - c.first_scan_time))/3600 as hours_parked,
            v.vessel_name, 
            v.vessel_type,
            ha.area_name as holding_area_name,
            c.stack_number,
            c.status
        FROM cars c
        JOIN scans s ON c.car_id = s.car_id
        JOIN users u ON s.worker_id = u.user_id
        LEFT JOIN vessels v ON c.vessel_id = v.vessel_id
        LEFT JOIN holding_areas ha ON c.holding_area_id = ha.holding_area_id
        WHERE c.date = %s AND c.is_in_holding = TRUE
    '''
    
    params = [date_filter]
    if shift:
        query += ' AND s.shift_number = %s'
        params.append(shift)
    
    query += ' ORDER BY c.first_scan_time'
    
    cur.execute(query, params)
    data = cur.fetchall()
    conn.close()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Holding Area"
    
    headers = ['Car ID', 'Vessel', 'Area', 'Unit', 'Worker', 'Time', 'Hours', 'Status']
    ws.append(headers)
    
    header_fill = PatternFill(start_color='f59e0b', end_color='f59e0b', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row in data:
        hours_parked = row['hours_parked'] or 0
        status_info = get_status_color(hours_parked)
        
        scan_time = row['last_scan_time'].strftime('%I:%M %p')
        vessel = f"{row['vessel_name']} ({row['vessel_type']})" if row['vessel_name'] else '-'
        
        ws.append([
            row['car_identifier'],
            vessel,
            row['holding_area_name'] or '-',
            row['stack_number'] or '-',
            row['worker_name'],
            scan_time,
            f"{hours_parked:.1f}h",
            f"{status_info['emoji']} {status_info['text']}"
        ])
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"holding_area_{date_filter}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True, 
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Not found'}), 404

@app.errorhandler(500)
def internal_error(e):
    print(f"Internal error: {e}")
    return jsonify({'error': 'Internal server error'}), 500

if __name__ == '__main__':  
    print("=" * 60)
    print("🚀 M Scanner - FIXED VERSION")
    print("=" * 60)
    print("📡 Server: http://localhost:5000")
    print("🔐 Login: admin / admin123")
    print("=" * 60)
    app.run(debug=True, host='0.0.0.0', port=5000)